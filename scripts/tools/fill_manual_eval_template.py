from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment

DEFAULT_RESPONSES_DIR = Path("artifacts/reports/manual_eval_responses")
DEFAULT_TEMPLATE_PATH = (
    DEFAULT_RESPONSES_DIR / "cartwise_manual_evaluation_template_v2.xlsx"
)
DEFAULT_OUTPUT_PATH = DEFAULT_RESPONSES_DIR / "cartwise_manual_evaluation_filled.xlsx"
SHEET_NAME = "Manual Evaluation"


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return ", ".join(str(item) for item in value if item is not None)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def format_price(value: Any) -> str:
    if value is None or value == "":
        return "Price unavailable"
    try:
        return f"{float(value):.2f}"
    except (TypeError, ValueError):
        return str(value)


def format_sources(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return str(value)


def format_evidence(
    evidence_items: Any, max_items: int = 3, max_chars: int = 500
) -> str:
    if not evidence_items:
        return ""

    lines: list[str] = []

    for index, item in enumerate(evidence_items[:max_items], start=1):
        if not isinstance(item, dict):
            lines.append(f"{index}. {normalize_text(item)}")
            continue

        rating = item.get("rating")
        score = item.get("score")
        review_id = item.get("review_id") or item.get("chunk_id") or ""
        text = item.get("chunk_text") or item.get("text") or ""

        text = normalize_text(text).replace("\n", " ").strip()
        if len(text) > max_chars:
            text = text[:max_chars].rstrip() + "..."

        meta_parts = []
        if rating is not None:
            meta_parts.append(f"rating={rating}")
        if score is not None:
            try:
                meta_parts.append(f"score={float(score):.4f}")
            except (TypeError, ValueError):
                meta_parts.append(f"score={score}")
        if review_id:
            meta_parts.append(f"id={review_id}")

        prefix = f"{index}. "
        if meta_parts:
            prefix += "[" + ", ".join(meta_parts) + "] "

        lines.append(prefix + text)

    return "\n".join(lines)


def find_header_row_and_columns(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        columns: dict[str, int] = {}
        for cell in ws[row_idx]:
            if cell.value is not None:
                columns[str(cell.value).strip()] = cell.column
        if "Query ID" in columns and "Rank" in columns:
            return row_idx, columns

    raise ValueError(
        "Cannot find header row. Expected columns include 'Query ID' and 'Rank'."
    )


def ensure_column(ws, header_row: int, columns: dict[str, int], name: str) -> int:
    if name in columns:
        return columns[name]

    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col).value = name
    columns[name] = new_col
    return new_col


def build_row_index(
    ws, header_row: int, columns: dict[str, int]
) -> dict[tuple[str, int], int]:
    query_col = columns["Query ID"]
    rank_col = columns["Rank"]

    row_index: dict[tuple[str, int], int] = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        query_id = ws.cell(row=row_idx, column=query_col).value
        rank = ws.cell(row=row_idx, column=rank_col).value

        if query_id is None or rank is None:
            continue

        try:
            rank_int = int(rank)
        except (TypeError, ValueError):
            continue

        row_index[(str(query_id).strip(), rank_int)] = row_idx

    return row_index


def set_cell(ws, row: int, col: int, value: Any, wrap: bool = False) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def fill_workbook(template_path: Path, responses_dir: Path, output_path: Path) -> None:
    if not template_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")

    if not responses_dir.exists():
        raise FileNotFoundError(f"Responses directory not found: {responses_dir}")

    wb = load_workbook(template_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_NAME}")

    ws = wb[SHEET_NAME]
    header_row, columns = find_header_row_and_columns(ws)

    required_columns = [
        "Product ID",
        "Product Title",
        "Brand",
        "Price",
        "Fusion Score",
        "Sources",
        "LLM Recommendation Reason",
        "LLM Potential Cons",
        "Retrieved Evidence Snippets",
        "API Latency ms",
        "Client Elapsed Seconds",
    ]

    for name in required_columns:
        ensure_column(ws, header_row, columns, name)

    row_index = build_row_index(ws, header_row, columns)

    filled_rows = 0
    missing_rows: list[str] = []
    error_files: list[str] = []

    json_files = sorted(
        path
        for path in responses_dir.glob("*.json")
        if not path.name.endswith(".error.json")
    )

    for json_path in json_files:
        data = load_json(json_path)

        if not data.get("success", False):
            error_files.append(json_path.name)
            continue

        query_id = str(data.get("query_id") or json_path.stem).strip()
        query = data.get("query", "")
        top_k = data.get("top_k", 3)
        client_elapsed = data.get("client_elapsed_seconds", "")
        response = data.get("response") or {}
        api_latency = response.get("latency_ms", "")
        results = response.get("results") or []

        for result in results:
            if not isinstance(result, dict):
                continue

            rank = result.get("rank")
            if rank is None:
                continue

            try:
                rank_int = int(rank)
            except (TypeError, ValueError):
                continue

            row = row_index.get((query_id, rank_int))
            if row is None:
                missing_rows.append(f"{query_id} rank={rank_int}")
                continue

            set_cell(ws, row, columns["Query"], query, wrap=True)
            set_cell(ws, row, columns["Top K"], top_k)
            set_cell(
                ws,
                row,
                columns["Product ID"],
                result.get("product_id") or result.get("parent_asin") or "",
            )
            set_cell(
                ws, row, columns["Product Title"], result.get("title") or "", wrap=True
            )
            set_cell(ws, row, columns["Brand"], result.get("brand") or "")
            set_cell(ws, row, columns["Price"], format_price(result.get("price")))
            set_cell(ws, row, columns["Fusion Score"], result.get("fusion_score"))
            set_cell(
                ws,
                row,
                columns["Sources"],
                format_sources(result.get("sources")),
                wrap=True,
            )
            set_cell(
                ws,
                row,
                columns["LLM Recommendation Reason"],
                result.get("reason") or "",
                wrap=True,
            )
            set_cell(
                ws,
                row,
                columns["LLM Potential Cons"],
                result.get("potential_cons") or "",
                wrap=True,
            )
            set_cell(
                ws,
                row,
                columns["Retrieved Evidence Snippets"],
                format_evidence(result.get("evidence")),
                wrap=True,
            )
            set_cell(ws, row, columns["API Latency ms"], api_latency)
            set_cell(ws, row, columns["Client Elapsed Seconds"], client_elapsed)

            filled_rows += 1

    text_heavy_columns = [
        "Query",
        "Product Title",
        "Sources",
        "LLM Recommendation Reason",
        "LLM Potential Cons",
        "Retrieved Evidence Snippets",
        "Notes",
    ]

    for name in text_heavy_columns:
        if name in columns:
            col_letter = ws.cell(row=header_row, column=columns[name]).column_letter
            if name == "Product Title":
                ws.column_dimensions[col_letter].width = 42
            elif name in {
                "LLM Recommendation Reason",
                "LLM Potential Cons",
                "Retrieved Evidence Snippets",
            }:
                ws.column_dimensions[col_letter].width = 45
            elif name == "Query":
                ws.column_dimensions[col_letter].width = 36
            else:
                ws.column_dimensions[col_letter].width = 24

    for row_idx in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 72

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Filled rows: {filled_rows}")
    print(f"Saved workbook: {output_path}")

    if missing_rows:
        print("Rows not found in template:")
        for item in missing_rows[:20]:
            print(f"  - {item}")
        if len(missing_rows) > 20:
            print(f"  ... and {len(missing_rows) - 20} more")

    if error_files:
        print("Skipped failed response files:")
        for item in error_files:
            print(f"  - {item}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--responses-dir", type=Path, default=DEFAULT_RESPONSES_DIR)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    args = parser.parse_args()

    fill_workbook(
        template_path=args.template,
        responses_dir=args.responses_dir,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()
